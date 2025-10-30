from django.shortcuts import render, redirect, get_object_or_404
from django.http import Http404, HttpResponse, JsonResponse
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy, reverse
from django.contrib import messages
from django.db import transaction
from django.db.models import Sum, Case, When, F, FloatField
from django.db.models.functions import Coalesce
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_protect
from utils.access import get_empresas_ids_for_cliente, get_clientes_ids_for_parceiro
from django.contrib.auth.mixins import LoginRequiredMixin
from accounts.decorators import cliente_can_view_lancamento, admin_required
from django.core.exceptions import ValidationError
from .models import Lancamentos, Anexos
from .forms import LancamentosForm, AnexosFormSet, LancamentoApprovalForm
from .permissions import LancamentoPermissionMixin, LancamentoClienteViewOnlyMixin, AdminRequiredMixin
## removido import duplicado de Http404/HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from django.utils.timezone import now, localtime
from empresas.models import Empresa
from django.utils.dateformat import format as date_format
from django.views.decorators.http import require_GET
from django.contrib.auth.decorators import login_required
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import permissions, status
from .serializers import LancamentoSerializer, AnexoSerializer
from .models import Lancamentos, Anexos
# --- Exportação de lançamentos para XLSX ---
from django.contrib.auth.decorators import login_required
@login_required
def exportar_lancamentos_xlsx(request):
    """Exporta lançamentos respeitando a mesma lógica de acesso da listagem:
    - Admin/Staff: todos
    - Cliente: união (empresas diretas + via sócio)
    - Parceiro: lançamentos de clientes vinculados à sua empresa_parceira
    """
    base = Lancamentos.objects.select_related(
        'id_adesao', 'id_adesao__cliente', 'id_adesao__cliente__id_company_vinculada'
    ).order_by('-data_criacao')
    user = request.user
    if user.is_superuser or user.is_staff:
        queryset = base
    elif hasattr(user, 'profile'):
        profile = user.profile
        if profile.eh_cliente:
            empresas_ids = get_empresas_ids_for_cliente(profile)
            queryset = base.filter(id_adesao__cliente__id_company_vinculada_id__in=empresas_ids) if empresas_ids else base.none()
        elif profile.eh_parceiro:
            clientes_ids = get_clientes_ids_for_parceiro(profile)
            queryset = base.filter(id_adesao__cliente__id_company_vinculada_id__in=clientes_ids) if clientes_ids else base.none()
        else:
            queryset = base.none()
    else:
        queryset = base.none()

    perdcomp = request.GET.get('perdcomp')
    if perdcomp:
        queryset = queryset.filter(id_adesao__perdcomp__icontains=perdcomp)
    aprovado = request.GET.get('aprovado')
    if aprovado == '1':
        queryset = queryset.filter(aprovado=True)
    elif aprovado == '0':
        queryset = queryset.filter(aprovado=False)

    # Criação do arquivo XLSX
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lançamentos"

    # Cabeçalho extra
    ws['A1'] = 'Relatório de Lançamentos'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f'Gerado em: {now().strftime("%d/%m/%Y %H:%M:%S")}'
    ws['A3'] = f'Usuário: {user.get_username()}'
    empresa_nome = getattr(getattr(user.profile, 'empresa_vinculada', None), 'razao_social', '-') if hasattr(user, 'profile') else '-'
    ws['A4'] = f'Empresa: {empresa_nome}'

    # Cabeçalho dos dados (linha 6)
    headers = [
        'PER/DCOMP Adesão',
        'Declaração PER/DCOMP',
        'Item',
        'Código da Guia',
        'Cliente',
        'Data',
        'Valor do Lançamento',
        'Sinal',
        'Saldo Restante',
        'Descrição',
        'Qtd. Anexos',
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = Font(bold=True)

    # Dados
    for idx, lanc in enumerate(queryset, start=7):
        ws.cell(row=idx, column=1, value=getattr(lanc.id_adesao, 'perdcomp', ''))
        ws.cell(row=idx, column=2, value=lanc.perdcomp_declaracao or '')
        ws.cell(row=idx, column=3, value=lanc.item or '')
        ws.cell(row=idx, column=4, value=lanc.codigo_guia or '')
        cliente = getattr(getattr(lanc.id_adesao, 'cliente', None), 'id_company_vinculada', None)
        cliente_nome = getattr(cliente, 'razao_social', str(cliente)) if cliente else ''
        ws.cell(row=idx, column=5, value=cliente_nome)
        ws.cell(row=idx, column=6, value=lanc.data_lancamento.strftime('%d/%m/%Y'))
        ws.cell(row=idx, column=7, value=lanc.valor)
        ws.cell(row=idx, column=8, value=lanc.sinal)
        ws.cell(row=idx, column=9, value=lanc.saldo_restante)
        ws.cell(row=idx, column=10, value=lanc.descricao if hasattr(lanc, 'descricao') else '')
        ws.cell(row=idx, column=11, value=lanc.anexos.count())

    # Ajuste de largura
    for col in range(1, len(headers)+1):
        ws.column_dimensions[get_column_letter(col)].width = 22

    # Resposta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="relatorio_lancamentos_{now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response)
    return response

class LancamentosListView(LancamentoClienteViewOnlyMixin, ListView):
    model = Lancamentos
    template_name = 'lancamentos_list.html'
    context_object_name = 'lancamentos'
    paginate_by = 10
    
    def _get_scoped_base_queryset(self):
        """Retorna queryset com escopo de acesso aplicado e filtro de perdcomp,
        mas sem filtrar por status de aprovação. Mantém ordenação e anotação de anexos.
        """
        from django.db.models import Count
        base = super().get_queryset().select_related(
            'id_adesao', 'id_adesao__cliente', 'id_adesao__cliente__id_company_vinculada'
        ).annotate(num_anexos=Count('anexos')).order_by('-data_criacao')
        user = self.request.user
        if user.is_superuser or user.is_staff:
            qs = base
        else:
            if not hasattr(user, 'profile'):
                return base.none()
            profile = user.profile
            if profile.eh_cliente:
                empresas_ids = get_empresas_ids_for_cliente(profile)
                if not empresas_ids:
                    return base.none()
                qs = base.filter(id_adesao__cliente__id_company_vinculada_id__in=empresas_ids)
            elif profile.eh_parceiro:
                clientes_ids = get_clientes_ids_for_parceiro(profile)
                if not clientes_ids:
                    return base.none()
                qs = base.filter(id_adesao__cliente__id_company_vinculada_id__in=clientes_ids)
            else:
                return base.none()

        perdcomp = self.request.GET.get('perdcomp')
        if perdcomp:
            qs = qs.filter(id_adesao__perdcomp__icontains=perdcomp)
        return qs

    def get_queryset(self):
        """Filtra lançamentos conforme escopo de acesso do usuário.
        Regras:
        - Admin/staff: tudo
        - Cliente: lançamentos de adesões cujas empresas estão em (empresas diretas + via sócio)
        - Parceiro: lançamentos de adesões de clientes vinculados à sua empresa_parceira
        """
        qs = self._get_scoped_base_queryset()
        aprovado = self.request.GET.get('aprovado')
        if aprovado == '1':
            qs = qs.filter(aprovado=True)
        elif aprovado == '0':
            qs = qs.filter(aprovado=False)
        return qs
        
    def get_context_data(self, **kwargs):
        """
        Adiciona parâmetros de filtro ao contexto para persistir a pesquisa na paginação.
        """
        context = super().get_context_data(**kwargs)
        context['current_filters'] = self.request.GET.dict()
        # Resumo sintético (ignora filtro de status, respeita escopo/perdcomp)
        qs_all = self._get_scoped_base_queryset()
        signed_expr = Case(
            When(sinal='-', then=Coalesce(F('valor'), 0.0) * -1),
            default=Coalesce(F('valor'), 0.0),
            output_field=FloatField()
        )
        aprovados_qs = qs_all.filter(aprovado=True)
        nao_aprov_qs = qs_all.filter(aprovado=False)
        aprovados_total = aprovados_qs.aggregate(total=Sum(signed_expr))['total'] or 0.0
        nao_aprov_total = nao_aprov_qs.aggregate(total=Sum(signed_expr))['total'] or 0.0
        context['resumo'] = {
            'aprovados': {
                'qtd': aprovados_qs.count(),
                'total': aprovados_total,
            },
            'nao_aprovados': {
                'qtd': nao_aprov_qs.count(),
                'total': nao_aprov_total,
            }
        }
        return context

class LancamentoDetailView(LoginRequiredMixin, DetailView):
    model = Lancamentos
    template_name = 'lancamentos_detail.html'
    context_object_name = 'lancamento'

    def get_queryset(self):
        base = super().get_queryset().select_related(
            'id_adesao', 'id_adesao__cliente', 'id_adesao__cliente__id_company_vinculada'
        )
        user = self.request.user
        if user.is_superuser or user.is_staff:
            return base
        if not hasattr(user, 'profile'):
            return base.none()
        profile = user.profile
        if profile.eh_cliente:
            empresas_ids = get_empresas_ids_for_cliente(profile)
            return base.filter(id_adesao__cliente__id_company_vinculada_id__in=empresas_ids) if empresas_ids else base.none()
        if profile.eh_parceiro:
            clientes_ids = get_clientes_ids_for_parceiro(profile)
            return base.filter(id_adesao__cliente__id_company_vinculada_id__in=clientes_ids) if clientes_ids else base.none()
        return base.none()

    def get(self, request, *args, **kwargs):
        pk = kwargs.get(self.pk_url_kwarg)
        queryset = self.get_queryset()
        try:
            self.object = queryset.get(pk=pk)
        except self.model.DoesNotExist:
            if self.model.objects.filter(pk=pk).exists():
                messages.error(request, "Você não tem permissão para visualizar este lançamento.")
                from django.shortcuts import render
                return render(request, 'forbidden.html', {
                    'message': "Você não tem permissão para visualizar este lançamento."
                }, status=403)
            raise Http404(f"Lançamento com ID {pk} não encontrado.")
        context = self.get_context_data(object=self.object)
        return self.render_to_response(context)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['anexos'] = self.object.anexos.all()
        return context
        
class LancamentoCreateView(AdminRequiredMixin, CreateView):
    model = Lancamentos
    form_class = LancamentosForm
    template_name = 'lancamentos_form.html'
    success_url = reverse_lazy('lancamentos:list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['anexos_formset'] = AnexosFormSet(self.request.POST, self.request.FILES)
        else:
            context['anexos_formset'] = AnexosFormSet()
        
        # Adicionar saldos das adesões para o JavaScript
        from adesao.models import Adesao
        import json
        adesoes_saldos = {}
        for adesao in Adesao.objects.all():
            adesoes_saldos[str(adesao.id)] = float(adesao.saldo_atual or 0)
        context['adesoes_saldos'] = json.dumps(adesoes_saldos)
        
        return context
    
    def form_valid(self, form):
        """Processa o formulário principal e o formset de anexos usando submissão tradicional."""
        context = self.get_context_data()
        anexos_formset = context['anexos_formset']
        try:
            with transaction.atomic():
                if not anexos_formset.is_valid():
                    return self.form_invalid(form)
                from django.core.exceptions import ValidationError as DjangoValidationError
                try:
                    self.object = form.save()
                except DjangoValidationError as ve:
                    # Propaga erros para o form
                    if hasattr(ve, 'message_dict'):
                        for field, msgs in ve.message_dict.items():
                            for m in msgs:
                                if field in form.fields:
                                    form.add_error(field, m)
                                else:
                                    form.add_error(None, m)
                    else:
                        msgs = []
                        if hasattr(ve, 'messages'):
                            msgs = ve.messages
                        elif hasattr(ve, 'message'):
                            msgs = [ve.message]
                        for m in msgs:
                            form.add_error(None, m)
                    return self.form_invalid(form)
                anexos_formset.instance = self.object
                anexos_formset.save()
                messages.success(self.request, 'Lançamento criado com sucesso!')
                return super().form_valid(form)
        except Exception as e:
            import logging
            logging.error(f"Erro ao salvar lançamento: {str(e)}")
            messages.error(self.request, f"Erro ao processar lançamento: {str(e)}")
            return self.form_invalid(form)
    
    def form_invalid(self, form):
        context = self.get_context_data(form=form)
        anexos_formset = context['anexos_formset']
        messages.error(self.request, 'Erro ao cadastrar lançamento. Verifique os campos.')
        if anexos_formset.errors:
            for i, form_errors in enumerate(anexos_formset.errors):
                if form_errors:
                    messages.error(self.request, f'Erro no anexo {i+1}: {form_errors}')
        if anexos_formset.non_form_errors():
            messages.error(self.request, f'Erros do formset: {anexos_formset.non_form_errors()}')
        return self.render_to_response(context)

# Removemos as views LancamentoUpdateView, LancamentoDeleteView e confirmar_lancamento
# já que lançamentos não podem mais ser editados ou excluídos

class AnexosUpdateView(AdminRequiredMixin, UpdateView):
    """View para editar apenas os anexos de um lançamento.
    
    Permite a edição de anexos, pois apenas os anexos são modificáveis
    após a criação do lançamento.
    """
    model = Lancamentos
    template_name = 'anexos_form.html'
    context_object_name = 'lancamento'
    
    def get_success_url(self):
        return reverse_lazy('lancamentos:detail', kwargs={'pk': self.object.pk})
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['anexos_formset'] = AnexosFormSet(
                self.request.POST, 
                self.request.FILES, 
                instance=self.object
            )
        else:
            context['anexos_formset'] = AnexosFormSet(instance=self.object)
        
        # Adiciona uma flag para indicar que pode editar anexos sempre
        context['pode_editar_anexos'] = True
        return context
    
    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        anexos_formset = AnexosFormSet(request.POST, request.FILES, instance=self.object)
        
        if anexos_formset.is_valid():
            with transaction.atomic():
                anexos_formset.save()
                messages.success(request, "Anexos atualizados com sucesso!")
                return redirect(self.get_success_url())
        else:
            return self.render_to_response(
                self.get_context_data(anexos_formset=anexos_formset)
            )


class LancamentoApprovalUpdateView(AdminRequiredMixin, UpdateView):
    model = Lancamentos
    form_class = LancamentoApprovalForm
    template_name = 'lancamentos_approval_form.html'

    def get_success_url(self):
        messages.success(self.request, 'Status de aprovação atualizado.')
        return reverse_lazy('lancamentos:detail', kwargs={'pk': self.object.pk})

    def get_queryset(self):
        base = super().get_queryset().select_related('id_adesao', 'id_adesao__cliente', 'id_adesao__cliente__id_company_vinculada')
        return base

    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        if self.object.aprovado:
            messages.info(request, 'Este lançamento já está aprovado e não pode ser alterado.')
            return redirect('lancamentos:detail', pk=self.object.pk)
        return super().dispatch(request, *args, **kwargs)

@login_required
@require_GET
def lancamento_history_json(request, pk):
    try:
        lanc = Lancamentos.objects.get(pk=pk)
    except Lancamentos.DoesNotExist:
        raise Http404
    manager = getattr(lanc, 'history', None) or getattr(lanc, 'historico', None)
    if manager is None:
        return JsonResponse({'error': 'Histórico não configurado.'}, status=400)

    def serialize_value(val):
        from django.db.models import Model
        import datetime
        if val is None:
            return None
        if isinstance(val, Model):
            return str(val)
        if isinstance(val, (datetime.datetime, datetime.date, datetime.time)):
            try:
                return val.isoformat()
            except Exception:
                return str(val)
        return val

    history = list(manager.all().order_by('history_date'))
    result = []
    for idx, record in enumerate(history):
        entry = {
            'id': record.id,
            'history_id': getattr(record, 'history_id', None),
            'history_date': date_format(localtime(record.history_date), 'd/m/Y H:i:s'),
            'history_user': getattr(record.history_user, 'username', None),
            'history_type': record.history_type,
            'changes': [],
            'note': ''
        }
        if idx == 0:
            for field in record._meta.fields:
                fname = field.name
                if fname in ('history_id','history_date','history_change_reason','history_type','history_user','id'):
                    continue
                val = getattr(record, fname, None)
                if val not in (None, ''):
                    entry['changes'].append({'field': fname, 'old': None, 'new': serialize_value(val)})
        else:
            prev = history[idx-1]
            try:
                diff = record.diff_against(prev)
                for c in diff.changes:
                    entry['changes'].append({'field': c.field, 'old': serialize_value(c.old), 'new': serialize_value(c.new)})
            except Exception:
                pass
            if not entry['changes']:
                ignore = {'history_id','history_date','history_change_reason','history_type','history_user','id'}
                manual = []
                for field in record._meta.fields:
                    fname = field.name
                    if fname in ignore:
                        continue
                    curr_val = getattr(record, fname, None)
                    prev_val = getattr(prev, fname, None)
                    if curr_val != prev_val:
                        manual.append({'field': fname, 'old': serialize_value(prev_val), 'new': serialize_value(curr_val)})
                if manual:
                    entry['changes'] = manual
                else:
                    entry['note'] = 'Alteração sem mudança perceptível.'
        result.append(entry)
    result.reverse()
    return JsonResponse({'object_id': lanc.id, 'history': result})


@login_required
@require_GET
def anexo_history_json(request, pk):
    try:
        anexo = Anexos.objects.get(pk=pk)
    except Anexos.DoesNotExist:
        raise Http404
    manager = getattr(anexo, 'history', None) or getattr(anexo, 'historico', None)
    if manager is None:
        return JsonResponse({'error': 'Histórico não configurado.'}, status=400)
    history = list(manager.all().order_by('history_date'))
    result = []
    for idx, record in enumerate(history):
        entry = {
            'id': record.id,
            'history_id': getattr(record, 'history_id', None),
            'history_date': date_format(localtime(record.history_date), 'd/m/Y H:i:s'),
            'history_user': getattr(record.history_user, 'username', None),
            'history_type': record.history_type,
            'changes': []
        }
        if idx == 0:
            for field in record._meta.fields:
                fname = field.name
                if fname in ('history_id','history_date','history_change_reason','history_type','history_user','id'):
                    continue
                val = getattr(record, fname, None)
                if val not in (None, ''):
                    entry['changes'].append({'field': fname, 'old': None, 'new': str(val)})
        else:
            prev = history[idx-1]
            try:
                diff = record.diff_against(prev)
                for c in diff.changes:
                    entry['changes'].append({'field': c.field, 'old': str(c.old), 'new': str(c.new)})
            except Exception:
                pass
        result.append(entry)
    result.reverse()
    return JsonResponse({'object_id': anexo.id, 'history': result})

# ================== DRF API ==================
class IsSuperAdmin(permissions.BasePermission):
    def has_permission(self, request, view):
        return request.user and request.user.is_authenticated and (request.user.is_superuser or request.user.is_staff)

class LancamentoListAPI(APIView):
    permission_classes = [permissions.IsAuthenticated, IsSuperAdmin]
    def get(self, request):
        objs = Lancamentos.objects.all().select_related('id_adesao')
        ser = LancamentoSerializer(objs, many=True)
        return Response(ser.data)

class LancamentoCreateAPI(APIView):
    permission_classes = [permissions.IsAuthenticated, IsSuperAdmin]
    def post(self, request):
        ser = LancamentoSerializer(data=request.data)
        if ser.is_valid():
            ser.save()
            return Response(ser.data, status=status.HTTP_201_CREATED)
        return Response(ser.errors, status=status.HTTP_400_BAD_REQUEST)

class LancamentoDetailAPI(APIView):
    permission_classes = [permissions.IsAuthenticated, IsSuperAdmin]
    def get_object(self, pk):
        return get_object_or_404(Lancamentos, pk=pk)
    def get(self, request, pk):
        ser = LancamentoSerializer(self.get_object(pk))
        return Response(ser.data)
    def patch(self, request, pk):
        obj = self.get_object(pk)
        ser = LancamentoSerializer(obj, data=request.data, partial=True)
        if ser.is_valid():
            ser.save()
            return Response(ser.data)
        return Response(ser.errors, status=status.HTTP_400_BAD_REQUEST)
    def delete(self, request, pk):
        obj = self.get_object(pk)
        obj.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

class AnexoListAPI(APIView):
    permission_classes = [permissions.IsAuthenticated, IsSuperAdmin]
    def get(self, request):
        objs = Anexos.objects.all()
        ser = AnexoSerializer(objs, many=True)
        return Response(ser.data)

class AnexoCreateAPI(APIView):
    permission_classes = [permissions.IsAuthenticated, IsSuperAdmin]
    def post(self, request):
        ser = AnexoSerializer(data=request.data)
        if ser.is_valid():
            ser.save()
            return Response(ser.data, status=status.HTTP_201_CREATED)
        return Response(ser.errors, status=status.HTTP_400_BAD_REQUEST)

class AnexoDetailAPI(APIView):
    permission_classes = [permissions.IsAuthenticated, IsSuperAdmin]
    def get_object(self, pk):
        return get_object_or_404(Anexos, pk=pk)
    def get(self, request, pk):
        ser = AnexoSerializer(self.get_object(pk))
        return Response(ser.data)
    def patch(self, request, pk):
        obj = self.get_object(pk)
        ser = AnexoSerializer(obj, data=request.data, partial=True)
        if ser.is_valid():
            ser.save()
            return Response(ser.data)
        return Response(ser.errors, status=status.HTTP_400_BAD_REQUEST)
    def delete(self, request, pk):
        obj = self.get_object(pk)
        obj.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


@login_required
@csrf_protect
@require_POST
def importar_recibo_lancamento(request):
    """
    Importa recibo de envio de declaração de compensação.
    Atualiza lançamento com número de controle, chave SERPRO e status='protocolado'.
    """
    from django.core.files.uploadedfile import UploadedFile
    from pdf import extract_text
    from utils.pdf_parser import parse_recibo_pedido_credito_text
    import tempfile
    import os
    import json
    from typing import Any
    
    pdf_file: UploadedFile | None = request.FILES.get('pdf')
    if not pdf_file:
        return JsonResponse({'ok': False, 'error': 'Arquivo PDF não enviado (campo "pdf").'}, status=400)
    
    from django.conf import settings
    status_code = 200
    response_payload: dict[str, Any] | None = None
    log_data: dict[str, Any] = {
        'user': getattr(request.user, 'username', None),
        'filename': pdf_file.name,
        'ts': timezone.now().isoformat(),
        'context': 'recibo_lancamento',
    }
    tmp_path = None

    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp:
            for chunk in pdf_file.chunks():
                tmp.write(chunk)
            tmp_path = tmp.name

        txt = extract_text(tmp_path) or ''
        parsed = parse_recibo_pedido_credito_text(txt)
        log_data['parsed'] = parsed.as_dict()

        numero_documento = (parsed.numero_documento or '').strip()
        if not numero_documento:
            status_code = 400
            response_payload = {'ok': False, 'error': 'Número do Documento não identificado no PDF.'}
        else:
            try:
                lancamento = Lancamentos.objects.get(perdcomp_declaracao=numero_documento)
            except Lancamentos.DoesNotExist:
                status_code = 404
                response_payload = {'ok': False, 'error': f'Lançamento com PER/DCOMP Declaração {numero_documento} não encontrado.'}

        if status_code == 200:
            missing = []
            numero_controle = (parsed.numero_controle or '').strip()
            if not numero_controle:
                missing.append('Número de Controle')
            autenticacao_serpro = (parsed.autenticacao_serpro or '').strip()
            if not autenticacao_serpro:
                missing.append('Autenticação SERPRO')
            if missing:
                status_code = 400
                response_payload = {'ok': False, 'error': f'Campo(s) não identificado(s) no PDF: {", ".join(missing)}.'}

        if status_code == 200:
            fields_to_update = []
            if numero_controle:
                lancamento.numero_controle = numero_controle
                fields_to_update.append('numero_controle')
            if autenticacao_serpro:
                lancamento.chave_seguranca_serpro = autenticacao_serpro
                fields_to_update.append('chave_seguranca_serpro')

            lancamento.status = 'protocolado'
            fields_to_update.append('status')
            lancamento.save(update_fields=fields_to_update)

            response_payload = {
                'ok': True,
                'created': False,
                'updated': True,
                'id': lancamento.pk,
                'detail_url': reverse('lancamentos:detail', kwargs={'pk': lancamento.pk}),
                'numero_controle': lancamento.numero_controle,
                'chave_seguranca_serpro': lancamento.chave_seguranca_serpro,
                'status': lancamento.get_status_display(),
                'numero_documento': numero_documento,
            }

    except Exception as e:
        status_code = 500
        response_payload = {'ok': False, 'error': f'Erro ao processar PDF: {str(e)}'}
    finally:
        try:
            log_data['status_code'] = status_code
            log_data['result'] = response_payload
            logs_dir = os.path.join(getattr(settings, 'MEDIA_ROOT', ''), 'import_logs')
            os.makedirs(logs_dir, exist_ok=True)
            safe_name = os.path.basename(pdf_file.name).replace(' ', '_')
            log_path = os.path.join(
                logs_dir,
                f"recibo_lanc_{timezone.now().strftime('%Y%m%d_%H%M%S')}_{safe_name}.json"
            )
            with open(log_path, 'w', encoding='utf-8') as fh:
                fh.write(json.dumps(log_data, ensure_ascii=False, indent=2))
        except Exception:
            pass
        try:
            if tmp_path and os.path.exists(tmp_path):
                os.remove(tmp_path)
        except Exception:
            pass

    return JsonResponse(response_payload or {'ok': False, 'error': 'Falha desconhecida.'}, status=status_code)
